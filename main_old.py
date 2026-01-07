"""
FastAPI File Converter Backend
Run with: uvicorn main:app --reload --port 8000
"""

import os
import base64
import tempfile
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse

# Conversion libraries
from pdf2docx import Converter as PDFToWordConverter
from docx import Document
from docx.shared import Inches
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from openpyxl import Workbook, load_workbook
import fitz  # PyMuPDF

app = FastAPI(
    title="File Converter API",
    description="API for converting files between PDF, Word, and Excel formats",
    version="1.0.0"
)

# CORS configuration
frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:5173")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[frontend_url, "http://localhost:5173", "http://localhost:8080", "*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def encode_file_to_base64(file_path: str) -> str:
    """Read a file and return its base64 encoded content."""
    with open(file_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


@app.get("/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "message": "File Converter API is running"}


@app.post("/api/convert/pdf-to-word")
async def convert_pdf_to_word(file: UploadFile = File(...)):
    """Convert PDF to Word document."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a PDF")
    
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save uploaded PDF
            pdf_path = os.path.join(temp_dir, "input.pdf")
            docx_path = os.path.join(temp_dir, "output.docx")
            
            content = await file.read()
            with open(pdf_path, "wb") as f:
                f.write(content)
            
            # Convert PDF to Word
            cv = PDFToWordConverter(pdf_path)
            cv.convert(docx_path)
            cv.close()
            
            # Return base64 encoded file
            output_filename = Path(file.filename).stem + ".docx"
            file_base64 = encode_file_to_base64(docx_path)
            
            return JSONResponse({
                "success": True,
                "filename": output_filename,
                "file": file_base64
            })
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/convert/word-to-pdf")
async def convert_word_to_pdf(file: UploadFile = File(...)):
    """Convert Word document to PDF."""
    if not file.filename.lower().endswith((".doc", ".docx")):
        raise HTTPException(status_code=400, detail="File must be a Word document")
    
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save uploaded Word file
            docx_path = os.path.join(temp_dir, "input.docx")
            pdf_path = os.path.join(temp_dir, "output.pdf")
            
            content = await file.read()
            with open(docx_path, "wb") as f:
                f.write(content)
            
            # Read Word document
            doc = Document(docx_path)
            
            # Create PDF
            c = canvas.Canvas(pdf_path, pagesize=letter)
            width, height = letter
            y_position = height - inch
            
            for para in doc.paragraphs:
                text = para.text
                if text.strip():
                    # Handle long lines by wrapping
                    words = text.split()
                    line = ""
                    for word in words:
                        test_line = line + " " + word if line else word
                        if c.stringWidth(test_line, "Helvetica", 12) < width - 2 * inch:
                            line = test_line
                        else:
                            c.drawString(inch, y_position, line)
                            y_position -= 14
                            line = word
                            
                            if y_position < inch:
                                c.showPage()
                                y_position = height - inch
                    
                    if line:
                        c.drawString(inch, y_position, line)
                        y_position -= 14
                    
                    y_position -= 6  # Paragraph spacing
                    
                    if y_position < inch:
                        c.showPage()
                        y_position = height - inch
            
            c.save()
            
            output_filename = Path(file.filename).stem + ".pdf"
            file_base64 = encode_file_to_base64(pdf_path)
            
            return JSONResponse({
                "success": True,
                "filename": output_filename,
                "file": file_base64
            })
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/convert/pdf-to-excel")
async def convert_pdf_to_excel(file: UploadFile = File(...)):
    """Extract tables from PDF to Excel."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a PDF")
    
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            pdf_path = os.path.join(temp_dir, "input.pdf")
            xlsx_path = os.path.join(temp_dir, "output.xlsx")
            
            content = await file.read()
            with open(pdf_path, "wb") as f:
                f.write(content)
            
            # Extract text from PDF using PyMuPDF
            doc = fitz.open(pdf_path)
            wb = Workbook()
            ws = wb.active
            ws.title = "Extracted Data"
            
            row = 1
            for page_num, page in enumerate(doc):
                text = page.get_text()
                lines = text.split('\n')
                
                # Add page header
                ws.cell(row=row, column=1, value=f"--- Page {page_num + 1} ---")
                row += 1
                
                for line in lines:
                    if line.strip():
                        # Try to split by common delimiters
                        parts = line.split('\t') if '\t' in line else line.split('  ')
                        parts = [p.strip() for p in parts if p.strip()]
                        
                        for col, part in enumerate(parts, 1):
                            ws.cell(row=row, column=col, value=part)
                        row += 1
                
                row += 1  # Empty row between pages
            
            doc.close()
            wb.save(xlsx_path)
            
            output_filename = Path(file.filename).stem + ".xlsx"
            file_base64 = encode_file_to_base64(xlsx_path)
            
            return JSONResponse({
                "success": True,
                "filename": output_filename,
                "file": file_base64
            })
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/convert/excel-to-pdf")
async def convert_excel_to_pdf(file: UploadFile = File(...)):
    """Convert Excel to PDF."""
    if not file.filename.lower().endswith((".xls", ".xlsx")):
        raise HTTPException(status_code=400, detail="File must be an Excel file")
    
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            xlsx_path = os.path.join(temp_dir, "input.xlsx")
            pdf_path = os.path.join(temp_dir, "output.pdf")
            
            content = await file.read()
            with open(xlsx_path, "wb") as f:
                f.write(content)
            
            # Read Excel file
            wb = load_workbook(xlsx_path)
            
            # Create PDF
            c = canvas.Canvas(pdf_path, pagesize=letter)
            width, height = letter
            
            for sheet in wb.worksheets:
                y_position = height - inch
                c.setFont("Helvetica-Bold", 14)
                c.drawString(inch, y_position, f"Sheet: {sheet.title}")
                y_position -= 24
                c.setFont("Helvetica", 10)
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        # Truncate if too long
                        if len(row_text) > 100:
                            row_text = row_text[:97] + "..."
                        c.drawString(inch, y_position, row_text)
                        y_position -= 14
                        
                        if y_position < inch:
                            c.showPage()
                            c.setFont("Helvetica", 10)
                            y_position = height - inch
                
                c.showPage()
            
            c.save()
            
            output_filename = Path(file.filename).stem + ".pdf"
            file_base64 = encode_file_to_base64(pdf_path)
            
            return JSONResponse({
                "success": True,
                "filename": output_filename,
                "file": file_base64
            })
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
