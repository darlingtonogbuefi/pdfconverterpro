# backend/services/pdf_to_excel.py

import pdfplumber  # Ensure pdfplumber is imported
from openpyxl import Workbook
from openpyxl.utils import get_column_letter  # Ensure get_column_letter is imported
from pathlib import Path
from backend.utils.file_utils import encode_file_to_base64
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def convert_pdf_to_excel(pdf_path: str, out_path: str, original_name: str):
    """
    Convert a PDF file to Excel using pdfplumber for both text and table extraction.
    Returns a dictionary containing success, filename, and base64-encoded file.
    """
    try:
        # Create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"Content_{Path(pdf_path).stem}"

        # Extract text and table data using pdfplumber
        text_content = extract_text_from_pdf(pdf_path)
        table_data = extract_tables_from_pdf(pdf_path)
        
        # Write extracted text content to Excel
        current_row = write_text_to_excel(ws, text_content)

        # Write extracted table data to Excel
        write_tables_to_excel(ws, table_data, current_row)

        # Save the Excel file to the output path
        wb.save(out_path)
        logging.info(f"Excel file saved to {out_path}")

        # Encode the output Excel file to base64 for frontend use
        encoded_file = encode_file_to_base64(out_path)

        return {
            "success": True,
            "filename": f"{Path(original_name).stem}.xlsx",
            "file": encoded_file
        }
    except Exception as e:
        logging.error(f"Error in convert_pdf_to_excel: {e}")
        return {
            "success": False,
            "message": f"An error occurred: {e}"
        }

def extract_text_from_pdf(pdf_path: str):
    """
    Extract text from the PDF using pdfplumber.
    The text will be placed into rows in the Excel file, trying to preserve the original layout.
    """
    text_content = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text_content += page.extract_text() or ""  # Avoid NoneType errors
    logging.info(f"Extracted {len(text_content)} characters of text from {pdf_path}")
    return text_content

def extract_tables_from_pdf(pdf_path: str):
    """
    Extract tables from all pages of a PDF file using pdfplumber.
    Each table is then mapped to Excel rows and columns.
    """
    table_data = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Extract tables from each page
            tables = page.extract_tables()

            if tables:
                logging.info(f"Found {len(tables)} table(s) on page {page.page_number}")
                table_data.extend(tables)
            else:
                logging.info(f"No tables found on page {page.page_number}")
    return table_data

def write_text_to_excel(ws, text_content):
    """
    Write extracted text content to the Excel worksheet while preserving layout.
    Each line of text from the PDF will be placed in a new row.
    """
    current_row = 1
    for line in text_content.split("\n"):
        if line.strip():  # Avoid empty lines
            ws.cell(row=current_row, column=1, value=line)
            current_row += 1
    ws.column_dimensions['A'].width = 50  # Adjust column width based on the content length
    logging.info(f"Written text content to Excel, {current_row - 1} rows")
    return current_row

def write_tables_to_excel(ws, table_data, starting_row):
    """
    Write extracted table data to the Excel worksheet, preserving the table structure.
    Tables from the PDF will be mapped to the corresponding rows and columns.
    """
    row_offset = starting_row
    for table_index, table in enumerate(table_data):
        for r, row in enumerate(table):
            for c, cell in enumerate(row):
                cell_value = cell if cell is not None else ""  # Handle None values gracefully
                ws.cell(row=row_offset + r, column=c + 1, value=cell_value)
        row_offset += len(table) + 2  # Leave a space between tables

    # Adjust column widths dynamically for each table based on the content length
    for col in range(1, len(table_data[0][0]) + 1):  # Iterate over columns
        max_length = 0
        for row in ws.iter_rows(min_row=starting_row, max_row=row_offset, min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width

    logging.info(f"Written {len(table_data)} tables to Excel")
